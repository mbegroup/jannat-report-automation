from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DAY_SHEET = re.compile(r"^\d{2}\.\d{2}$")

METRIC_LABELS = {
    "rooms_sold": "Продано номеров",
    "occupancy": "Загрузка %",
    "adr": "ADR (средний тариф НФ)",
    "revpar": "RevPAR",
    "trevpar": "TrevPAR",
    "b2c": "Номера B2C (прямые",
    "b2b": "Номера B2B (корп.",
    "online": "Онлайн (Booking",
    "rooms_revenue": "ИТОГО НФ",
    "b2c_rooms": "Номеров B2C (шт)",
    "b2b_rooms": "Номеров B2B (шт)",
    "online_rooms": "Номеров Онлайн (шт)",
    "conference": "Конференции / аренда залов",
    "restaurant": "Ресторан — алякарт",
    "pool_bar": "Пул-бар",
    "banquet": "Банкет / Фуршет",
    "sbt": "СБТ (бассейн",
    "medical": "Медицинский центр",
    "rent": "Коммерческая аренда",
    "other_income": "Прочие доходы (минибар",
    "income": "ИТОГО ДОХОДОВ",
    "income_net": "ИТОГО ДОХОДОВ без налога",
    "variable_expenses": "ИТОГО ПЕРЕМЕННЫХ РАСХОДОВ",
    "fixed_expenses": "ИТОГО ПОСТОЯННЫХ РАСХОДОВ",
    "operating_expenses": "ИТОГО ОПЕРАЦИОННЫХ РАСХОДОВ",
    "ebitda": "EBITDA",
    "depreciation": "Амортизация по всем продуктам",
    "credit": "Кредитные обязательства",
    "ebit": "EBIT (прибыль до налога)",
    "profit_tax": "Налог на прибыль",
    "net_profit": "ЧИСТАЯ ПРИБЫЛЬ",
    "profitability": "РЕНТАБЕЛЬНОСТЬ",
    "memberships": "Активных абонементов",
}

METRIC_ALIASES = {
    "operating_expenses": ("ИТОГО ОПЕРАЦИОННЫХ РАСХОДОВ", "ИТОГО ВСЕХ РАСХОДОВ"),
    "ebitda": ("EBITDA", "ОПЕРАЦИОННАЯ ПРИБЫЛЬ (Daily GOP)"),
}

OBJECT_PATTERNS = (
    ("sb", ("SUL_BSH", "SULTAN_BSH", "СУЛТАН БИШКЕК")),
    ("skk", ("SUL_KK", "SULTAN_KK", "СУЛТАН КЫЗЫЛ")),
    ("jrkt", ("JRKT", "КОЙ-ТАШ")),
    ("jrja", ("JRJA", "ЖАЛАЛ-АБАД")),
    ("jrb", ("JRB", "РЕЖЕНСИ БИШКЕК")),
    ("jro", ("JRO", "РЕЗОРТ ОШ")),
)

OPTIONAL_METRICS = {
    "conference", "restaurant", "pool_bar", "banquet", "sbt", "medical",
    "rent", "other_income", "depreciation", "credit", "profit_tax",
    "memberships", "online_rooms", "income_net", "depreciation", "credit",
    "ebit", "profit_tax", "net_profit", "profitability",
}


def _num(value: Any) -> float:
    if value in (None, "", "-") or str(value).strip().casefold() in {"нет", "н/д", "n/a"}:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _clean(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        if value.startswith("[") and value.endswith("]"):
            value = value[1:-1].strip()
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _metric(sheet, row: int | None) -> dict[str, Any]:
    if row is None:
        return {
            "label": "", "unit": "", "plan_day": 0.0, "plan_month": 0.0,
            "plan_accum": 0.0, "fact_day": 0.0, "fact_accum": 0.0,
            "day_pct": 0.0, "accum_pct": 0.0, "comment": "",
            "recommendation": "",
        }
    return {
        "label": _clean(sheet[f"B{row}"].value),
        "unit": _clean(sheet[f"C{row}"].value),
        "plan_day": _num(sheet[f"D{row}"].value),
        "plan_month": _num(sheet[f"E{row}"].value),
        "plan_accum": _num(sheet[f"F{row}"].value),
        "fact_day": _num(sheet[f"G{row}"].value),
        "fact_accum": _num(sheet[f"H{row}"].value),
        "day_pct": _num(sheet[f"I{row}"].value),
        "accum_pct": _num(sheet[f"J{row}"].value),
        "comment": _clean(sheet[f"K{row}"].value),
        "recommendation": _clean(sheet[f"L{row}"].value),
    }


def _close(actual: float, expected: float, tolerance: float = 2.0) -> bool:
    return abs(actual - expected) <= tolerance


def _find_row(sheet, label: str) -> int:
    needle = label.casefold()
    for row in range(1, sheet.max_row + 1):
        for column in (1, 2, 3):
            value = str(sheet.cell(row, column).value or "").casefold()
            if needle in value:
                return row
    raise ValueError(f"Не найден показатель: {label}")


def _metric_rows(sheet) -> dict[str, int | None]:
    rows: dict[str, int | None] = {}
    # More specific labels must win over their shorter parents.
    ordered = sorted(METRIC_LABELS.items(), key=lambda item: len(item[1]), reverse=True)
    for key, label in ordered:
        labels = METRIC_ALIASES.get(key, (label,))
        rows[key] = None
        for candidate in labels:
            try:
                rows[key] = _find_row(sheet, candidate)
                break
            except ValueError:
                continue
        if rows[key] is None:
            if key not in OPTIONAL_METRICS:
                raise ValueError(f"Не найден показатель: {label}")
    return rows


def _object_key(path: Path, hotel: str) -> str | None:
    haystack = f"{path.name} {hotel}".upper()
    for key, patterns in OBJECT_PATTERNS:
        if any(pattern in haystack for pattern in patterns):
            return key
    return None


def _room_fund(sheet) -> int:
    value = " ".join(str(sheet.cell(5, col).value or "") for col in range(1, 5))
    match = re.search(r"(\d+)\s*ном", value, re.IGNORECASE)
    return int(match.group(1)) if match else 0


def _placeholder(value: Any) -> bool:
    text = str(value or "").strip()
    return not text or (text.startswith("[") and text.endswith("]"))


def _filled_day(sheet, rows: dict[str, int | None]) -> bool:
    daily_keys = (
        "rooms_sold", "b2c", "b2b", "online", "conference", "restaurant",
        "pool_bar", "banquet", "sbt", "medical", "rent", "other_income",
        "variable_expenses",
    )
    if any(rows[key] and _num(sheet.cell(rows[key], 7).value) != 0 for key in daily_keys):
        return True
    for label in ("Главная проблема", "Что предпринято", "Требует решения", "TOP-3 задачи", "Идея / предложение"):
        try:
            row = _find_row(sheet, label)
        except ValueError:
            continue
        if not _placeholder(sheet.cell(row, 3).value):
            return True
    return False


def parse_gm_august_report(
    path: str | Path,
    target_sheet: str | None = None,
    workbook: Any | None = None,
) -> dict[str, Any]:
    path = Path(path)
    book = workbook or load_workbook(path, data_only=True)
    day_names = [name for name in book.sheetnames if DAY_SHEET.fullmatch(name)]
    if not day_names:
        raise ValueError("В книге не найдены дневные листы формата ДД.ММ")

    day_names.sort(key=lambda value: tuple(reversed(tuple(map(int, value.split("."))))))
    reference_rows = _metric_rows(book[day_names[0]])
    filled = [name for name in day_names if _filled_day(book[name], reference_rows)]
    if not filled:
        raise ValueError("В книге нет ни одного дня с заполненной строкой ИТОГО ДОХОДОВ")
    latest_name = target_sheet or filled[-1]
    if latest_name not in filled:
        raise ValueError(f"Лист {latest_name} не заполнен в файле {path.name}")
    sheet = book[latest_name]
    rows = _metric_rows(sheet)
    metrics = {key: _metric(sheet, row) for key, row in rows.items()}

    income_parts = ["rooms_revenue", "conference", "restaurant", "pool_bar", "banquet", "sbt", "medical", "rent", "other_income"]
    income_day_sum = sum(metrics[key]["fact_day"] for key in income_parts)
    income_accum_sum = sum(metrics[key]["fact_accum"] for key in income_parts)
    checks = [
        {
            "name": "Доходы дня сходятся с детализацией",
            "ok": _close(metrics["income"]["fact_day"], income_day_sum),
            "expected": round(income_day_sum),
            "actual": round(metrics["income"]["fact_day"]),
        },
        {
            "name": "Накопленные доходы сходятся с детализацией",
            "ok": _close(metrics["income"]["fact_accum"], income_accum_sum),
            "expected": round(income_accum_sum),
            "actual": round(metrics["income"]["fact_accum"]),
        },
        {
            "name": "Операционные расходы = переменные + постоянные",
            "ok": _close(
                metrics["operating_expenses"]["fact_accum"],
                metrics["variable_expenses"]["fact_accum"] + metrics["fixed_expenses"]["fact_accum"],
            ),
            "expected": round(metrics["variable_expenses"]["fact_accum"] + metrics["fixed_expenses"]["fact_accum"]),
            "actual": round(metrics["operating_expenses"]["fact_accum"]),
        },
        {
            "name": "EBITDA = доходы − операционные расходы",
            "ok": _close(
                metrics["ebitda"]["fact_accum"],
                metrics["income"]["fact_accum"] - metrics["operating_expenses"]["fact_accum"],
            ),
            "expected": round(metrics["income"]["fact_accum"] - metrics["operating_expenses"]["fact_accum"]),
            "actual": round(metrics["ebitda"]["fact_accum"]),
        },
    ]
    if rows["net_profit"] and rows["ebit"]:
        checks.append({
            "name": "Чистая прибыль = EBIT − налог",
            "ok": _close(
                metrics["net_profit"]["fact_accum"],
                metrics["ebit"]["fact_accum"] - metrics["profit_tax"]["fact_accum"],
            ),
            "expected": round(metrics["ebit"]["fact_accum"] - metrics["profit_tax"]["fact_accum"]),
            "actual": round(metrics["net_profit"]["fact_accum"]),
        })

    reviews = []
    review_header = _find_row(sheet, "ОТЗЫВЫ ГОСТЕЙ")
    for row in range(review_header + 2, min(review_header + 7, sheet.max_row + 1)):
        reviews.append({
            "platform": _clean(sheet[f"B{row}"].value),
            "rating": _clean(sheet[f"C{row}"].value),
            "new_reviews": _clean(sheet[f"D{row}"].value),
            "comment": _clean(sheet[f"F{row}"].value),
        })

    return {
        "report_type": "gm_august",
        "source_file": path.name,
        "object_key": _object_key(path, str(sheet["C2"].value or "")),
        "meta": {
            "hotel": _clean(sheet["C2"].value),
            "date": _clean(sheet["C3"].value),
            "gm": _clean(sheet["C4"].value),
            "room_fund": _room_fund(sheet),
            "latest_sheet": latest_name,
            "filled_days": len(filled),
            "day_sheets": len(day_names),
            "missing_before_latest": [name for name in day_names[: day_names.index(latest_name) + 1] if name not in filled],
        },
        "metrics": metrics,
        "checks": checks,
        "verified": all(check["ok"] for check in checks),
        "operational": {
            "incident": _clean(sheet.cell(_find_row(sheet, "Главная проблема"), 3).value),
            "action": _clean(sheet.cell(_find_row(sheet, "Что предпринято"), 3).value),
            "escalation": _clean(sheet.cell(_find_row(sheet, "Требует решения"), 3).value),
            "tomorrow": _clean(sheet.cell(_find_row(sheet, "TOP-3 задачи"), 3).value),
            "idea": _clean(sheet.cell(_find_row(sheet, "Идея / предложение"), 3).value),
        },
        "reviews": reviews,
        "filled_sheet_names": filled,
    }
