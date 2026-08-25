from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _clean(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _num(value: Any) -> float:
    return 0.0 if value in (None, "") else float(value)


def _sheet(book, token: str):
    for sheet in book.worksheets:
        if token.lower() in sheet.title.lower():
            return sheet
    raise ValueError(f"Не найден лист: {token}")


def parse_commercial_report(path: str | Path) -> dict[str, Any]:
    path = Path(path)
    book = load_workbook(path, data_only=True)
    formulas = load_workbook(path, data_only=False)
    sheet = _sheet(book, "Отчёт отдела продаж")
    formula_sheet = formulas[sheet.title]
    online_sheet = _sheet(book, "Онлайн")

    managers = []
    for row in range(14, 19):
        managers.append({
            "name": _clean(sheet[f"A{row}"].value),
            "calls": int(_num(sheet[f"B{row}"].value)),
            "meetings": int(_num(sheet[f"C{row}"].value)),
            "companies": int(_num(sheet[f"D{row}"].value)),
            "contracts": int(_num(sheet[f"E{row}"].value)),
            "revenue": round(_num(sheet[f"F{row}"].value)),
            "focus": _clean(sheet[f"G{row}"].value),
            "comment": _clean(sheet[f"K{row}"].value),
        })

    detail = {
        "calls_all": sum(x["calls"] for x in managers),
        "calls_b2b": sum(x["calls"] for x in managers[:4]),
        "meetings": sum(x["meetings"] for x in managers),
        "companies": sum(x["companies"] for x in managers),
        "contracts": sum(x["contracts"] for x in managers),
        "revenue": sum(x["revenue"] for x in managers),
    }
    reported = {
        "calls_all": round(_num(sheet["B19"].value)),
        "meetings": round(_num(sheet["C19"].value)),
        "companies": round(_num(sheet["D19"].value)),
        "contracts": round(_num(sheet["E19"].value)),
        "revenue": round(_num(sheet["F19"].value)),
    }
    kpi = {
        "calls": {"plan_day": round(_num(sheet["C8"].value)), "fact_day": round(_num(sheet["D8"].value)), "plan_month": round(_num(sheet["F8"].value)), "fact_month": round(_num(sheet["G8"].value))},
        "meetings": {"plan_day": round(_num(sheet["C9"].value)), "fact_day": round(_num(sheet["D9"].value)), "plan_month": round(_num(sheet["F9"].value)), "fact_month": round(_num(sheet["G9"].value))},
        "revenue": {"plan_day": round(_num(sheet["C10"].value)), "fact_day": round(_num(sheet["D10"].value)), "plan_month": round(_num(sheet["F10"].value)), "fact_month": round(_num(sheet["G10"].value))},
    }
    checks = [
        {"name": "Строка ИТОГО менеджеров", "ok": all(abs(detail[k] - reported[k]) <= 1 for k in reported), "expected": detail, "actual": reported},
        {"name": "KPI звонков B2B", "ok": detail["calls_b2b"] == kpi["calls"]["fact_day"], "expected": detail["calls_b2b"], "actual": kpi["calls"]["fact_day"]},
        {"name": "KPI встреч", "ok": detail["meetings"] == kpi["meetings"]["fact_day"], "expected": detail["meetings"], "actual": kpi["meetings"]["fact_day"]},
        {"name": "KPI выручки дня", "ok": abs(detail["revenue"] - kpi["revenue"]["fact_day"]) <= 1, "expected": detail["revenue"], "actual": kpi["revenue"]["fact_day"]},
    ]
    pipeline = []
    for row in range(23, 25):
        if sheet[f"A{row}"].value:
            pipeline.append({"client": _clean(sheet[f"A{row}"].value), "object": _clean(sheet[f"B{row}"].value), "period": _clean(sheet[f"D{row}"].value), "amount": round(_num(sheet[f"F{row}"].value)), "stage": _clean(sheet[f"G{row}"].value), "next_step": _clean(sheet[f"K{row}"].value)})
    online_filled = any(online_sheet.cell(r, c).value not in (None, "") for r in range(6, 11) for c in range(2, 8))
    return {
        "source_file": path.name,
        "meta": {"date": _clean(sheet["B4"].value), "weekday": _clean(sheet["D4"].value), "leader": _clean(sheet["F4"].value), "staff": _clean(sheet["H4"].value)},
        "kpi": kpi, "managers": managers, "detail_totals": detail, "reported_totals": reported,
        "checks": checks, "verified": all(x["ok"] for x in checks), "pipeline": pipeline,
        "pipeline_total": sum(x["amount"] for x in pipeline), "online_filled": online_filled,
        "texts": {"result": _clean(sheet["A27"].value), "problem": _clean(sheet["A29"].value), "action": _clean(sheet["A31"].value), "tomorrow": _clean(sheet["A33"].value), "idea": _clean(sheet["A35"].value)},
        "formula_audit": {"D8": formula_sheet["D8"].value, "D9": formula_sheet["D9"].value, "D10": formula_sheet["D10"].value},
    }

